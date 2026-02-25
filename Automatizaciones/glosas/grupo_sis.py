# Automatizacon/glosas/grupo_sis.py

import os
import re
from openpyxl import load_workbook
from Core.api_gema import query_api_gema

def limpiar_texto(texto):
    """
    Elimina caracteres que no son válidos en XML 1.0 (causan IllegalCharacterError en openpyxl).
    """
    if not isinstance(texto, str):
        return texto
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', texto)

def consultar_items_de_glosa_api(factura: str, estado: str, progreso_callback=None) -> list[dict]:
    try:
        match = re.match(r'([A-Za-z]+)(\d+)', factura)
        if not match:
            return []
        prefijo, num = match.groups()

        sql_gl_docn = f"gl_docn FROM [gema10.d/salud/datos/glo_cab] WHERE fc_serie = '{prefijo}' AND fc_docn = {num} ORDER BY gl_fecha DESC"
        if progreso_callback:
            progreso_callback(f"  [API GEMA] Consultando ID para {factura}...")
        
        resultados_cab = query_api_gema(sql_gl_docn)
        if not resultados_cab:
            return []

        gl_docn = resultados_cab[0]['gl_docn']
        sql_gl_det = f"codigo, vr_glosa, motivo_res FROM [gema10.d/salud/datos/glo_det] WHERE gl_docn = {gl_docn} AND estatus1 = '{estado}'"
        if progreso_callback:
            progreso_callback(f"  [API GEMA] Obteniendo detalles de ítems (gl_docn: {gl_docn}, estado: {estado})...")
        
        return query_api_gema(sql_gl_det)
    except Exception as e:
        if progreso_callback:
            progreso_callback(f"  - ERROR CRÍTICO durante la consulta a la API: {e}")
        return []

def procesar_glosas_grupo_sis(ruta_excel: str, lista_glosas_texto: str, progreso_callback):
    """
    Procesa glosas para Grupo SIS integrando la lógica de procesador_glosas_excel.py.
    """
    # 1. Analizar lista de glosas desde texto
    glosas_a_procesar = []
    lineas = lista_glosas_texto.strip().split('\n')
    for linea in lineas:
        partes = linea.strip().split()
        if len(partes) == 2:
            glosas_a_procesar.append((partes[0].upper(), partes[1].upper()))
        else:
            progreso_callback(f"  -> Advertencia: Formato incorrecto, se omitirá: '{linea}'")

    if not glosas_a_procesar:
        progreso_callback("❌ No hay glosas válidas para procesar.")
        return 0, 0, []

    # 2. Cargar Excel
    try:
        progreso_callback(f"\n🔄 Abriendo archivo: {ruta_excel}...")
        wb = load_workbook(ruta_excel)
        ws = wb.active
        progreso_callback("Lectura exitosa.")
    except Exception as e:
        progreso_callback(f"❌ ERROR FATAL al abrir el Excel: {e}")
        return 0, 0, []

    # 3. Mapeo de columnas
    col_map = {
        "Factura": None,
        "Valor Glosa Tarifa": None,
        "Valor Aceptado": None,
        "Valor No Aceptado": None,
        "Observaciones": None,
        "Glosa Factura": None
    }

    encabezados = {str(celda.value).strip(): celda.column for celda in ws[1] if celda.value}
    for k in col_map.keys():
        col_map[k] = encabezados.get(k)

    missing_cols = [k for k, v in col_map.items() if v is None]
    if missing_cols:
        progreso_callback(f"❌ ERROR: Faltan columnas en el Excel: {missing_cols}")
        return 0, 0, []

    exitos = 0
    fallos = 0
    total_filas_actualizadas = 0
    reporte_final = []

    # 4. Procesamiento
    for factura, estado in glosas_a_procesar:
        progreso_callback(f"\n--- Procesando {factura} | Estado: {estado} ---")
        items_api = consultar_items_de_glosa_api(factura, estado, progreso_callback)
        
        if not items_api:
            progreso_callback("  - Sin ítems en la API o factura no encontrada.")
            fallos += 1
            reporte_final.append(f"{factura} (Sin datos API)")
            continue

        # Buscar filas en Excel
        filas_factura = []
        for i in range(2, ws.max_row + 1):
            val = ws.cell(i, col_map['Factura']).value
            valor_excel = str(val).strip().upper() if val else ""
            if valor_excel == factura:
                filas_factura.append(i)

        if not filas_factura:
            progreso_callback("  - No se encontró la factura en el Excel.")
            fallos += 1
            reporte_final.append(f"{factura} (No en Excel)")
            continue

        # Aplicar lógica de glosas
        total_api = sum(int(float(item.get('vr_glosa', 0))) for item in items_api)
        total_excel_rows = sum(int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0) for i in filas_factura)
        
        # Estrategia 1: Suma
        if total_excel_rows > 0 and abs(total_api - total_excel_rows) < 5:
            progreso_callback(f"  -> Coincidencia por Suma ({total_api}). Aplicando a todas las filas.")
            motivo_api = items_api[0].get('motivo_res', '')
            for i in filas_factura:
                v_tarifa = int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0)
                if estado == 'AI':
                    ws.cell(i, col_map['Valor Aceptado'], v_tarifa)
                    ws.cell(i, col_map['Valor No Aceptado'], 0)
                else:
                    ws.cell(i, col_map['Valor Aceptado'], 0)
                    ws.cell(i, col_map['Valor No Aceptado'], v_tarifa)
                ws.cell(i, col_map['Observaciones'], limpiar_texto(motivo_api))
                total_filas_actualizadas += 1
        else:
            # Estrategia 2: Item a Item
            progreso_callback(f"  -> Suma no coincide. Usando coincidencia exacta por ítem.")
            for item in items_api:
                v_api = int(float(item.get('vr_glosa', 0)))
                motivo_api = item.get('motivo_res', '')
                for i in filas_factura:
                    v_excel = int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0)
                    if v_excel == v_api and not ws.cell(i, col_map['Valor Aceptado']).value and not ws.cell(i, col_map['Valor No Aceptado']).value:
                        if estado == 'AI':
                            ws.cell(i, col_map['Valor Aceptado'], v_api)
                            ws.cell(i, col_map['Valor No Aceptado'], 0)
                        else:
                            ws.cell(i, col_map['Valor Aceptado'], 0)
                            ws.cell(i, col_map['Valor No Aceptado'], v_api)
                        ws.cell(i, col_map['Observaciones'], limpiar_texto(motivo_api))
                        total_filas_actualizadas += 1
                        break

        # Validación Final de la Factura
        suma_aceptado = sum(int(ws.cell(i, col_map['Valor Aceptado']).value or 0) for i in filas_factura)
        suma_no_aceptado = sum(int(ws.cell(i, col_map['Valor No Aceptado']).value or 0) for i in filas_factura)
        total_calc = suma_aceptado + suma_no_aceptado
        val_gf = ws.cell(filas_factura[0], col_map['Glosa Factura']).value
        val_gf_num = int(val_gf) if val_gf is not None else 0

        if total_calc == val_gf_num and all(int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0) == (int(ws.cell(i, col_map['Valor Aceptado']).value or 0) + int(ws.cell(i, col_map['Valor No Aceptado']).value or 0)) for i in filas_factura):
            exitos += 1
            reporte_final.append(f"{factura} (OK)")
        else:
            fallos += 1
            reporte_final.append(f"{factura} (Inconsistente)")
            progreso_callback(f"  ❌ Error de validación en {factura}")

    # Guardar
    try:
        wb.save(ruta_excel)
        progreso_callback(f"\n💾 Cambios guardados en {ruta_excel}")
    except Exception as e:
        progreso_callback(f"❌ ERROR al guardar Excel: {e}")

    return exitos, fallos, reporte_final
