import os
import re
from openpyxl import load_workbook, Workbook
from Core.api_gema import query_api_gema

def extraer_prefijo_numero(factura):
    """Separa el prefijo (letras) del número de la factura."""
    match = re.match(r"([A-Za-z]+)(\d+)", str(factura).strip())
    if match:
        return match.group(1).upper(), match.group(2)
    return None, None

def consultar_tipo_glosa():
    print("="*60)
    print("CONSULTA DE TIPO DE GLOSA (GEMA) - MODO REFINADO")
    print("="*60)

    # 1. Obtener inputs
    gr_docn = input("\nPaso 1A: Ingresa el número de Cuenta de Cobro (gr_docn): ").strip()
    ruta_input = input("Paso 1B: Arrastra el archivo Excel con las facturas y presiona Enter: ").strip().strip('"').strip("'")
    
    if not gr_docn.isdigit():
        print(f"❌ Error: El número de cuenta '{gr_docn}' debe ser puramente numérico.")
        return

    if not os.path.exists(ruta_input):
        print(f"❌ Error: El archivo no existe: {ruta_input}")
        return

    # 2. Cargar Excel y extraer facturas del usuario
    try:
        wb_in = load_workbook(ruta_input, data_only=True)
        ws_in = wb_in.active
        
        facturas_excel = []
        for i, row in enumerate(ws_in.iter_rows(min_row=1, max_col=1, values_only=True)):
            valor = row[0]
            if not valor: continue
            
            valor_str = str(valor).strip()
            # Omitir encabezado
            if i == 0 and valor_str.upper() in ["FACTURA", "CODIGO", "FACTURAS", "INVOICE"]:
                continue
                
            facturas_excel.append(valor_str)
        
        if not facturas_excel:
            print("❌ No se encontraron facturas válidas en el Excel.")
            return
        print(f"✅ Se cargaron {len(facturas_excel)} facturas del Excel.")
    except Exception as e:
        print(f"❌ Error leyendo el Excel: {e}")
        return

    # 3. Obtener Glosas vinculadas a la Cuenta de Cobro (glo_red)
    print(f"\nPaso 2: Buscando glosas asociadas a la Cuenta {gr_docn} en glo_red...")
    try:
        # glo_red vincula gr_docn con gl_docn
        sql_red = f"gl_docn FROM [gema10.d/salud/datos/glo_red] WHERE gr_docn = {gr_docn}"
        data_red = query_api_gema(sql_red)
        
        if not data_red:
            print(f"⚠️ No se encontraron glosas vinculadas a la cuenta {gr_docn}.")
            return
            
        # Extraer gl_docn únicos (pueden repetirse por items en glo_red)
        glosas_ids = list(set([str(int(item['gl_docn'])) for item in data_red if item.get('gl_docn')]))
        print(f"✅ Se identificaron {len(glosas_ids)} glosas vinculadas a la cuenta.")
        
    except Exception as e:
        print(f"❌ Error en consulta glo_red: {e}")
        return

    # 4. Obtener detalles de factura de esas glosas (glo_cab)
    print(f"\nPaso 3: Trayendo detalles de facturas (serie, número, tipo) desde glo_cab...")
    mapa_facturas_validas = {} # (serie, numero) -> tipo
    
    # Procesamos glosas una a una (o podrías hacer batches si el driver lo permite, pero individual es más seguro)
    # Mostramos contador por ser muchas glosas
    for idx, gl_id in enumerate(glosas_ids):
        try:
            # glo_cab vincula gl_docn con fc_serie y fc_docn
            sql_cab = f"tipo, fc_serie, fc_docn FROM [gema10.d/salud/datos/glo_cab] WHERE gl_docn = {gl_id}"
            data_cab = query_api_gema(sql_cab)
            
            for row in data_cab:
                serie = str(row.get('fc_serie')).strip().upper()
                numero = str(int(row.get('fc_docn'))).strip()
                tipo = str(row.get('tipo')).strip() if row.get('tipo') else "N/A"
                
                # Guardamos en nuestro mapa de validación
                mapa_facturas_validas[(serie, numero)] = tipo
                
            if (idx + 1) % 50 == 0:
                print(f"   ... procesadas {idx + 1} de {len(glosas_ids)} glosas.")
                
        except Exception as e:
            print(f"  ⚠️ Error consultando glosa {gl_id}: {e}")

    print(f"✅ Mapa de facturas de la cuenta construido. ({len(mapa_facturas_validas)} facturas encontradas)")

    # 5. Cruzar con el Excel y generar resultados
    print("\nPaso 4: Cruzando datos con tu lista de Excel...")
    resultados_finales = []
    for factura in facturas_excel:
        prefijo, numero = extraer_prefijo_numero(factura)
        
        if not prefijo or not numero:
            resultados_finales.append({"Factura": factura, "Tipo": "ERROR_FORMATO"})
            continue
            
        # Buscamos en nuestro mapa de la cuenta
        tipo_encontrado = mapa_facturas_validas.get((prefijo, numero))
        
        if tipo_encontrado:
            print(f"  🔹 {factura} -> {tipo_encontrado}")
            resultados_finales.append({"Factura": factura, "Tipo": tipo_encontrado})
        else:
            print(f"  ⛔ {factura} -> NO EN CUENTA {gr_docn}")
            resultados_finales.append({"Factura": factura, "Tipo": "NO_EN_CUENTA"})

    # 6. Guardar en Excel
    try:
        output_name = "resultados_tipos_glosa.xlsx"
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Resultados"
        ws_out.append(["Factura Cruzada", "Tipo Identificado"])
        
        for res in resultados_finales:
            ws_out.append([res["Factura"], res["Tipo"]])
            
        wb_out.save(output_name)
        print(f"\n✅ PROCESO FINALIZADO.")
        print(f"📄 Archivo generado: {os.path.abspath(output_name)}")
    except Exception as e:
        print(f"❌ Error guardando resultados: {e}")

if __name__ == "__main__":
    consultar_tipo_glosa()
