import os
import re
from openpyxl import load_workbook, Workbook
from Core.api_gema import query_api_gema

def extraer_prefijo_numero(factura):
    """Separa el prefijo (letras) del número de la factura."""
    match = re.match(r"([A-Za-z]+)(\d+)", str(factura).strip())
    if match:
        return match.group(1).upper(), match.group(2)
    return None, None

def consultar_tipo_glosa():
    print("="*60)
    print("CONSULTA DE TIPO DE GLOSA (GEMA) - MODO EXCEL DIRECTO")
    print("="*60)

    # 1. Obtener inputs
    ruta_input = input("\nArrastra el archivo Excel con las facturas y presiona Enter: ").strip().strip('"').strip("'")
    
    if not os.path.exists(ruta_input):
        print(f"❌ Error: El archivo no existe: {ruta_input}")
        return

    # 2. Cargar Excel y extraer facturas
    try:
        wb_in = load_workbook(ruta_input, data_only=True)
        ws_in = wb_in.active
        
        facturas_excel = []
        for i, row in enumerate(ws_in.iter_rows(min_row=1, max_col=1, values_only=True)):
            valor = row[0]
            if not valor: continue
            
            valor_str = str(valor).strip()
            # Omitir encabezado si existe
            if i == 0 and valor_str.upper() in ["FACTURA", "CODIGO", "FACTURAS", "INVOICE"]:
                continue
                
            facturas_excel.append(valor_str)
        
        if not facturas_excel:
            print("❌ No se encontraron facturas válidas en el Excel.")
            return
        print(f"✅ Se cargaron {len(facturas_excel)} facturas del Excel.")
    except Exception as e:
        print(f"❌ Error leyendo el Excel: {e}")
        return

    # 3. Consultar GEMA para cada factura
    print(f"\nProcesando {len(facturas_excel)} facturas...")
    resultados_finales = []
    
    for idx, factura in enumerate(facturas_excel):
        prefijo, numero = extraer_prefijo_numero(factura)
        
        if not prefijo or not numero:
            print(f"  ⚠️ {factura} -> ERROR FORMATO")
            resultados_finales.append({"Factura": factura, "Tipo": "ERROR_FORMATO"})
            continue
            
        try:
            # Consultamos la glosa más reciente para esta factura
            # Usamos freg DESC para obtener siempre el último registro
            sql = f"tipo FROM [gema10.d/salud/datos/glo_cab] WHERE fc_serie = '{prefijo}' AND fc_docn = {numero} ORDER BY freg DESC"
            data = query_api_gema(sql)
            
            if data and len(data) > 0:
                tipo_encontrado = str(data[0].get('tipo', 'S/T')).strip()
                print(f"  [{idx+1}/{len(facturas_excel)}] {factura} -> {tipo_encontrado}")
                resultados_finales.append({"Factura": factura, "Tipo": tipo_encontrado})
            else:
                print(f"  [{idx+1}/{len(facturas_excel)}] {factura} -> NO ENCONTRADA")
                resultados_finales.append({"Factura": factura, "Tipo": "NO_ENCONTRADA"})
                
        except Exception as e:
            print(f"  ⚠️ Error consultando {factura}: {e}")
            resultados_finales.append({"Factura": factura, "Tipo": "ERROR_API"})

    # 4. Guardar en Excel
    try:
        output_name = "resultados_tipos_glosa.xlsx"
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Resultados"
        ws_out.append(["Factura Original", "Tipo Identificado"])
        
        for res in resultados_finales:
            ws_out.append([res["Factura"], res["Tipo"]])
            
        wb_out.save(output_name)
        print(f"\n✅ PROCESO FINALIZADO.")
        print(f"📄 Archivo generado: {os.path.abspath(output_name)}")
    except Exception as e:
        print(f"❌ Error guardando resultados: {e}")

if __name__ == "__main__":
    consultar_tipo_glosa()
