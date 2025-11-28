# procesador_glosas_recons.py

import os
import re
from openpyxl import load_workbook

# ======================================================================
# IMPORTACIÓN REAL
# ======================================================================
try:
    from Core.api_gema import query_api_gema
except ImportError:
    print("ERROR FATAL: No se pudo encontrar el archivo 'Core/api_gema.py'.")
    exit()


# ======================================================================
# FUNCIONES DE SOPORTE
# ======================================================================
def obtener_ruta_excel() -> str:
    while True:
        ruta = input("Paso 1: Arrastra tu archivo Excel (.xlsx) aquí y presiona Enter: ").strip().replace("'", "").replace('"', '')
        if os.path.exists(ruta) and ruta.lower().endswith('.xlsx'):
            return ruta
        print("Ruta no válida. Asegúrate de que el archivo existe y es un .xlsx.")


def obtener_glosas_desde_input() -> list[tuple[str, str]]:
    print("\n" + "="*70)
    print("Paso 2: Pega tu lista de glosas o reconsideraciones. Formato: Factura ESTADO")
    print("\nEjemplo: COEX14393 C2")
    print("\nPresiona Enter en una línea vacía para finalizar.")
    print("="*70)
    glosas_a_procesar = []
    while True:
        linea = input().strip()
        if not linea:
            break
        partes = linea.split()
        if len(partes) == 2:
            glosas_a_procesar.append((partes[0].upper(), partes[1].upper()))
        else:
            print(f"  -> Advertencia: Formato incorrecto, se omitirá: '{linea}'")
    if glosas_a_procesar:
        print(f"\nOK. Se procesarán {len(glosas_a_procesar)} registros.")
    return glosas_a_procesar


def consultar_items_de_glosa_api(factura: str, estado: str) -> list[dict]:
    try:
        match = re.match(r'([A-Za-z]+)(\d+)', factura)
        if not match:
            return []
        prefijo, num = match.groups()

        sql_gl_docn = f"gl_docn FROM [gema10.d/salud/datos/glo_cab] WHERE fc_serie = '{prefijo}' AND fc_docn = {num} ORDER BY gl_fecha DESC"
        print(f"  [API GEMA] Consultando ID para {factura}...")
        resultados_cab = query_api_gema(sql_gl_docn)
        if not resultados_cab:
            return []

        gl_docn = resultados_cab[0]['gl_docn']
        sql_gl_det = f"codigo, vr_glosa, motivo_res FROM [gema10.d/salud/datos/glo_det] WHERE gl_docn = {gl_docn} AND estatus1 = '{estado}'"
        print(f"  [API GEMA] Obteniendo detalles de ítems (gl_docn: {gl_docn}, estado: {estado})...")
        return query_api_gema(sql_gl_det)
    except Exception as e:
        print(f"  - ERROR CRÍTICO durante la consulta a la API: {e}")
        return []


# ======================================================================
# PROCESAMIENTO PRINCIPAL
# ======================================================================
def procesar_archivo(tipo: str):
    ruta_excel = obtener_ruta_excel()
    glosas_a_procesar = obtener_glosas_desde_input()
    if not glosas_a_procesar:
        return

    try:
        print(f"\n🔄 Abriendo archivo: {ruta_excel}...")
        wb = load_workbook(ruta_excel)
        ws = wb.active
        print("Lectura exitosa.")
    except Exception as e:
        print(f"ERROR FATAL al abrir el Excel: {e}")
        return

    # ==================================================================
    # CONFIGURACIÓN SEGÚN TIPO
    # ==================================================================
    # ==================================================================
    # CONFIGURACIÓN SEGÚN TIPO
    # ==================================================================
    if tipo == "glosas":
        col_map = {
            "Factura": None,
            "Valor Glosa Tarifa": None,
            "Valor Aceptado": None,
            "Valor No Aceptado": None,
            "Observaciones": None,
            "Glosa Factura": None         # Campo único total
        }
    else:  # reconsideraciones / devoluciones
        col_map = {
            "Factura": None,
            "Valor Objetado": None,
            "Aceptado(1:Si/0:No)": None,
            "Observaciones": None
        }

    encabezados = {str(celda.value).strip(): celda.column for celda in ws[1] if celda.value}
    for k in col_map.keys():
        col_map[k] = encabezados.get(k)

    # Validar columnas obligatorias
    missing_cols = [k for k, v in col_map.items() if v is None]
    if missing_cols:
        print(f"❌ ERROR: Faltan columnas requeridas en el Excel: {missing_cols}")
        print(f"Columnas detectadas: {list(encabezados.keys())}")
        return

    reporte = {'exitos': set(), 'advertencias': set(), 'fallos': set()}
    total_items_actualizados = 0

    # ==================================================================
    # PROCESAMIENTO DE CADA FACTURA
    # ==================================================================
    for factura, estado in glosas_a_procesar:
        print(f"\n--- Procesando {factura} | Estado: {estado} ---")
        items_api = consultar_items_de_glosa_api(factura, estado)
        
        # Si la API devuelve vacío o mensaje especial, manejarlo
        if not items_api:
            print("  - Sin ítems en la API o factura no encontrada.")
            # Aquí podríamos marcar como REVISAR si así se desea, pero por ahora reportamos fallo
            reporte['fallos'].add(f"{factura} (Sin datos API)")
            continue

        filas_factura = []
        for i in range(2, ws.max_row + 1):
            valor_factura = str(ws.cell(i, col_map['Factura']).value).strip().upper() if ws.cell(i, col_map['Factura']).value else ""
            if valor_factura == factura:
                filas_factura.append(i)

        if not filas_factura:
            print("  - No se encontró la factura en el Excel (FALLO).")
            reporte['fallos'].add(f"{factura} (No en Excel)")
            continue

        items_encontrados_count = 0

        # ==============================================================
        # MODO GLOSAS
        # ==============================================================
        if tipo == "glosas":
            # Calcular totales para estrategia de coincidencia
            total_api = sum(int(float(item.get('vr_glosa', 0))) for item in items_api)
            total_excel_rows = 0
            for i in filas_factura:
                total_excel_rows += int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0)
            
            match_por_suma = False
            
            # ESTRATEGIA 1: Coincidencia por Suma Total (Prioritaria)
            # Si el total de la API coincide con el total de las filas del Excel,
            # asumimos que la respuesta de la API aplica para TODAS las filas.
            if total_excel_rows > 0 and abs(total_api - total_excel_rows) < 5: # Tolerancia de 5 pesos
                print(f"  -> Coincidencia por Suma detectada ({total_api}). Aplicando a todas las filas.")
                match_por_suma = True
                
                # Usamos la observación del primer ítem (o concatenamos si son distintas)
                motivo_api = items_api[0].get('motivo_res', '')
                
                for i in filas_factura:
                    valor_tarifa = int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0)
                    
                    if estado == 'AI':
                        ws.cell(i, col_map['Valor Aceptado'], valor_tarifa)
                        ws.cell(i, col_map['Valor No Aceptado'], 0)
                    else:
                        ws.cell(i, col_map['Valor Aceptado'], 0)
                        ws.cell(i, col_map['Valor No Aceptado'], valor_tarifa)
                    
                    ws.cell(i, col_map['Observaciones'], motivo_api)
                    items_encontrados_count += 1
                    total_items_actualizados += 1

            # ESTRATEGIA 2: Coincidencia Item a Item (Fallback)
            # Si la suma no cuadra, intentamos casar valores exactos individualmente.
            else:
                print(f"  -> Suma no coincide (API:{total_api} vs Excel:{total_excel_rows}). Usando coincidencia exacta por ítem.")
                for item in items_api:
                    valor_api = int(float(item.get('vr_glosa', 0)))
                    motivo_api = item.get('motivo_res', '')
                    
                    # Buscar fila que coincida con el valor de la glosa
                    for i in filas_factura:
                        valor_excel = int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0)
                        
                        # Solo llenar si coincide valor y aún no ha sido llenado
                        v_aceptado_actual = ws.cell(i, col_map['Valor Aceptado']).value
                        v_no_aceptado_actual = ws.cell(i, col_map['Valor No Aceptado']).value
                        
                        if valor_excel == valor_api and (not v_aceptado_actual and not v_no_aceptado_actual):
                            if estado == 'AI':
                                ws.cell(i, col_map['Valor Aceptado'], valor_api)
                                ws.cell(i, col_map['Valor No Aceptado'], 0)
                            else:
                                ws.cell(i, col_map['Valor Aceptado'], 0)
                                ws.cell(i, col_map['Valor No Aceptado'], valor_api)
                            
                            ws.cell(i, col_map['Observaciones'], motivo_api)
                            items_encontrados_count += 1
                            total_items_actualizados += 1
                            break # Pasar al siguiente item de la API
            
            # 2. Validación Estricta Post-Procesamiento
            es_valida = True
            mensaje_error = ""
            
            suma_aceptado = 0
            suma_no_aceptado = 0
            valor_glosa_factura_unico = None
            
            # Obtener Valor Glosa Factura (debe ser único para la factura)
            val_gf = ws.cell(filas_factura[0], col_map['Glosa Factura']).value
            valor_glosa_factura_unico = int(val_gf) if val_gf is not None else 0

            for i in filas_factura:
                v_tarifa = int(ws.cell(i, col_map['Valor Glosa Tarifa']).value or 0)
                v_aceptado = int(ws.cell(i, col_map['Valor Aceptado']).value or 0)
                v_no_aceptado = int(ws.cell(i, col_map['Valor No Aceptado']).value or 0)
                
                suma_aceptado += v_aceptado
                suma_no_aceptado += v_no_aceptado
                
                # Validación Fila: Tarifa == Aceptado + No Aceptado
                if v_tarifa != (v_aceptado + v_no_aceptado):
                    es_valida = False
                    mensaje_error = f"Fila {i} inconsistente: Tarifa({v_tarifa}) != Acept({v_aceptado}) + NoAcept({v_no_aceptado})"
                    break
            
            # Validación Factura: Glosa Factura == Suma Total
            if es_valida:
                total_calculado = suma_aceptado + suma_no_aceptado
                if valor_glosa_factura_unico != total_calculado:
                    es_valida = False
                    mensaje_error = f"Total inconsistente: Glosa Factura({valor_glosa_factura_unico}) != Suma({total_calculado})"

            # 3. Clasificación y Reporte
            if es_valida:
                # Determinar Estado Consolidado (RAD o REVISAR) para el reporte
                estado_final = "REVISAR"
                if suma_aceptado >= valor_glosa_factura_unico:
                    estado_final = "RAD"
                
                reporte['exitos'].add(f"{factura} ({estado_final})")
            else:
                reporte['fallos'].add(f"{factura} - {mensaje_error}")
                print(f"  ❌ ERROR VALIDACIÓN {factura}: {mensaje_error}")

        # ==============================================================
        # MODO RECONSIDERACIONES / DEVOLUCIONES
        # ==============================================================
        else:
            for i in filas_factura:
                ws.cell(i, col_map['Aceptado(1:Si/0:No)'], 1 if estado == 'AI' else 0)
                # Si hay múltiples ítems, usa el primer motivo_res disponible
                motivo = items_api[0].get('motivo_res', '') if items_api else ''
                ws.cell(i, col_map['Observaciones'], motivo)
                items_encontrados_count += 1
                total_items_actualizados += 1
            
            if items_encontrados_count > 0:
                reporte['exitos'].add(factura)
            else:
                reporte['fallos'].add(factura)

    # ==================================================================
    # GUARDADO Y REPORTE FINAL
    # ==================================================================
    try:
        wb.save(ruta_excel)
        print(f"\n💾 Cambios guardados directamente en {ruta_excel}")
        print(f"✔ Total de filas actualizadas: {total_items_actualizados}")
    except Exception as e:
        print(f"\nERROR al guardar el archivo Excel: {e}")

    print("\n" + "="*80)
    print("REPORTE FINAL")
    print("="*80)
    print(f"✅ Éxitos: {len(reporte['exitos'])}")
    if reporte['exitos']:
        for f in sorted(reporte['exitos']):
            print(f"   - {f}")
    print(f"❌ Fallos (Sin Conciliación): {len(reporte['fallos'])}")
    if reporte['fallos']:
        for f in sorted(reporte['fallos']):
            print(f"   - {f}")
    print("="*80)


# ======================================================================
# PUNTO DE ENTRADA
# ======================================================================
def main():
    print("="*70)
    print("Seleccione el modo de procesamiento:")
    print("1. Glosas")
    print("2. Reconsideraciones / Devoluciones")
    print("="*70)
    opcion = input("Opción (1/2): ").strip()

    if opcion == "1":
        procesar_archivo("glosas")
    elif opcion == "2":
        procesar_archivo("recons")
    else:
        print("Opción inválida. Saliendo.")


if __name__ == "__main__":
    main()
